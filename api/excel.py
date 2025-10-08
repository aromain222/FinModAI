#!/usr/bin/env python3
"""
Excel Generator - Banker-Style Workbook Creation
Generates professional Excel files for DCF, LBO, Trading Comps, and Merger models
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import xlsxwriter

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Generates banker-style Excel workbooks"""
    
    def __init__(self):
        self.output_dir = "generated_models"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Define styles
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.input_font = Font(name="Calibri", size=10, color="0000FF")
        self.formula_font = Font(name="Calibri", size=10, color="000000")
        self.warning_font = Font(name="Calibri", size=10, color="FF0000")
        self.warning_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        self.center_alignment = Alignment(horizontal="center", vertical="center")
    
    def generate_workbook(self, model_type: str, assumptions_result, preview: Dict[str, Any], filename: str) -> str:
        """Generate Excel workbook for the specified model"""
        file_path = os.path.join(self.output_dir, filename)
        
        if model_type == "dcf":
            self._create_dcf_workbook(file_path, assumptions_result, preview)
        elif model_type == "lbo":
            self._create_lbo_workbook(file_path, assumptions_result, preview)
        elif model_type == "comps":
            self._create_comps_workbook(file_path, assumptions_result, preview)
        elif model_type == "merger":
            self._create_merger_workbook(file_path, assumptions_result, preview)
        else:
            raise ValueError(f"Unknown model type: {model_type}")
        
        logger.info(f"Generated Excel workbook", file_path=file_path, model_type=model_type)
        return file_path
    
    def _create_dcf_workbook(self, file_path: str, assumptions_result, preview: Dict[str, Any]):
        """Create DCF model workbook"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_assumptions_sheet(wb, assumptions_result)
        self._create_historicals_sheet(wb, assumptions_result)
        self._create_dcf_sheet(wb, assumptions_result, preview)
        self._create_summary_sheet(wb, assumptions_result, preview, "DCF")
        
        wb.save(file_path)
    
    def _create_lbo_workbook(self, file_path: str, assumptions_result, preview: Dict[str, Any]):
        """Create LBO model workbook"""
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_assumptions_sheet(wb, assumptions_result)
        self._create_historicals_sheet(wb, assumptions_result)
        self._create_lbo_sheet(wb, assumptions_result, preview)
        self._create_summary_sheet(wb, assumptions_result, preview, "LBO")
        
        wb.save(file_path)
    
    def _create_comps_workbook(self, file_path: str, assumptions_result, preview: Dict[str, Any]):
        """Create Trading Comps workbook"""
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_assumptions_sheet(wb, assumptions_result)
        self._create_historicals_sheet(wb, assumptions_result)
        self._create_comps_sheet(wb, assumptions_result, preview)
        self._create_summary_sheet(wb, assumptions_result, preview, "Trading Comps")
        
        wb.save(file_path)
    
    def _create_merger_workbook(self, file_path: str, assumptions_result, preview: Dict[str, Any]):
        """Create Merger model workbook"""
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_assumptions_sheet(wb, assumptions_result)
        self._create_historicals_sheet(wb, assumptions_result)
        self._create_merger_sheet(wb, assumptions_result, preview)
        self._create_summary_sheet(wb, assumptions_result, preview, "Merger")
        
        wb.save(file_path)
    
    def _create_assumptions_sheet(self, wb: Workbook, assumptions_result):
        """Create assumptions sheet"""
        ws = wb.create_sheet("Assumptions")
        
        # Headers
        headers = ["Metric", "Value", "Source", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment
        
        # Assumptions data
        assumptions = assumptions_result.assumptions
        wacc = assumptions_result.wacc
        terminal = assumptions_result.terminal
        
        data = [
            ["Revenue Growth (Y1)", f"{assumptions.revenue_growth[0]:.1%}", "Derived", "Based on historical CAGR"],
            ["Revenue Growth (Y5)", f"{assumptions.revenue_growth[4]:.1%}", "Derived", "Faded to terminal growth"],
            ["Operating Margin", f"{assumptions.operating_margin[0]:.1%}", "Derived", "3-year average"],
            ["Tax Rate", f"{assumptions.tax_rate:.1%}", "Derived", "Effective tax rate"],
            ["WACC", f"{wacc.wacc:.1%}", "Calculated", f"Ke={wacc.ke:.1%}, Kd={wacc.kd_after:.1%}"],
            ["Terminal Growth", f"{terminal.g:.1%}", "Assumed", "Perpetuity growth rate"],
            ["Beta", f"{wacc.beta:.2f}", "Market", "From provider"],
            ["Risk-Free Rate", f"{wacc.rf:.1%}", "FRED", "10Y Treasury"],
            ["Equity Risk Premium", f"{wacc.erp:.1%}", "Assumed", "Market ERP"],
        ]
        
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.thin_border
                if col == 2:  # Value column
                    cell.font = self.input_font
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_historicals_sheet(self, wb: Workbook, assumptions_result):
        """Create historicals sheet"""
        ws = wb.create_sheet("Historicals")
        
        historicals = assumptions_result.historicals
        
        # Headers
        headers = ["Year", "Revenue", "EBITDA", "Op Margin", "D&A", "CapEx", "ΔNWC", "Tax Rate"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment
        
        # Historical data
        for row, year in enumerate(historicals.years, 2):
            ws.cell(row=row, column=1, value=year).border = self.thin_border
            ws.cell(row=row, column=2, value=historicals.revenue[row-2]).border = self.thin_border
            ws.cell(row=row, column=3, value=historicals.operating_income[row-2]).border = self.thin_border
            ws.cell(row=row, column=4, value=historicals.op_margin[row-2]).border = self.thin_border
            ws.cell(row=row, column=5, value=historicals.da[row-2]).border = self.thin_border
            ws.cell(row=row, column=6, value=historicals.capex[row-2]).border = self.thin_border
            ws.cell(row=row, column=7, value=historicals.delta_nwc[row-2]).border = self.thin_border
            
            tax_rate = historicals.tax_expense[row-2] / historicals.pretax_income[row-2] if historicals.pretax_income[row-2] > 0 else 0
            ws.cell(row=row, column=8, value=tax_rate).border = self.thin_border
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_dcf_sheet(self, wb: Workbook, assumptions_result, preview: Dict[str, Any]):
        """Create DCF model sheet"""
        ws = wb.create_sheet("DCF Model")
        
        # Headers
        headers = ["Year", "Revenue", "EBITDA", "UFCF", "PV UFCF", "Cumulative PV"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment
        
        # DCF projections
        years = preview["years"]
        revenue = preview["revenue"]
        ebitda = preview["ebitda"]
        ufcf = preview["ufcf"]
        pv_ufcf = preview["pv_ufcf"]
        
        cumulative_pv = 0
        for row, year in enumerate(years, 2):
            cumulative_pv += pv_ufcf[row-2]
            
            ws.cell(row=row, column=1, value=year).border = self.thin_border
            ws.cell(row=row, column=2, value=revenue[row-2]).border = self.thin_border
            ws.cell(row=row, column=3, value=ebitda[row-2]).border = self.thin_border
            ws.cell(row=row, column=4, value=ufcf[row-2]).border = self.thin_border
            ws.cell(row=row, column=5, value=pv_ufcf[row-2]).border = self.thin_border
            ws.cell(row=row, column=6, value=cumulative_pv).border = self.thin_border
        
        # Terminal value section
        terminal_row = len(years) + 4
        ws.cell(row=terminal_row, column=1, value="Terminal Value").font = Font(bold=True)
        ws.cell(row=terminal_row, column=2, value=preview["terminal_value"]).border = self.thin_border
        
        ws.cell(row=terminal_row+1, column=1, value="PV Terminal Value").font = Font(bold=True)
        ws.cell(row=terminal_row+1, column=2, value=preview["pv_terminal"]).border = self.thin_border
        
        # Valuation summary
        summary_row = terminal_row + 3
        ws.cell(row=summary_row, column=1, value="Enterprise Value").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=preview["ev"]).border = self.thin_border
        
        ws.cell(row=summary_row+1, column=1, value="Net Debt").font = Font(bold=True)
        ws.cell(row=summary_row+1, column=2, value=preview["net_debt"]).border = self.thin_border
        
        ws.cell(row=summary_row+2, column=1, value="Equity Value").font = Font(bold=True)
        ws.cell(row=summary_row+2, column=2, value=preview["equity_value"]).border = self.thin_border
        
        ws.cell(row=summary_row+3, column=1, value="Implied Price").font = Font(bold=True)
        ws.cell(row=summary_row+3, column=2, value=preview["implied_price"]).border = self.thin_border
        
        ws.cell(row=summary_row+4, column=1, value="Upside %").font = Font(bold=True)
        cell = ws.cell(row=summary_row+4, column=2, value=preview["upside_pct"])
        cell.border = self.thin_border
        if preview["upside_pct"] > 0:
            cell.font = Font(color="00FF00", bold=True)  # Green for positive
        else:
            cell.font = Font(color="FF0000", bold=True)  # Red for negative
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_lbo_sheet(self, wb: Workbook, assumptions_result, preview: Dict[str, Any]):
        """Create LBO model sheet"""
        ws = wb.create_sheet("LBO Model")
        
        # Sources & Uses
        ws.cell(row=1, column=1, value="Sources & Uses").font = Font(size=14, bold=True)
        
        sources_uses = preview["sources_uses"]
        
        # Sources
        ws.cell(row=3, column=1, value="SOURCES").font = Font(bold=True)
        row = 4
        for source, amount in sources_uses["sources"].items():
            ws.cell(row=row, column=1, value=source.replace("_", " ").title())
            ws.cell(row=row, column=2, value=amount)
            row += 1
        
        # Uses
        ws.cell(row=3, column=4, value="USES").font = Font(bold=True)
        row = 4
        for use, amount in sources_uses["uses"].items():
            ws.cell(row=row, column=4, value=use.replace("_", " ").title())
            ws.cell(row=row, column=5, value=amount)
            row += 1
        
        # Debt Stack
        debt_start_row = 10
        ws.cell(row=debt_start_row, column=1, value="Debt Stack").font = Font(size=14, bold=True)
        
        debt_headers = ["Tranche", "Amount", "Rate", "Amortization"]
        for col, header in enumerate(debt_headers, 1):
            cell = ws.cell(row=debt_start_row+2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        debt_stack = preview["debt_stack"]
        for row, tranche in enumerate(debt_stack, debt_start_row+3):
            ws.cell(row=row, column=1, value=tranche["name"]).border = self.thin_border
            ws.cell(row=row, column=2, value=tranche["amount"]).border = self.thin_border
            ws.cell(row=row, column=3, value=tranche["rate"]).border = self.thin_border
            ws.cell(row=row, column=4, value=tranche["amortization"]).border = self.thin_border
        
        # Returns Summary
        returns_row = debt_start_row + len(debt_stack) + 5
        ws.cell(row=returns_row, column=1, value="Returns Summary").font = Font(size=14, bold=True)
        
        ws.cell(row=returns_row+2, column=1, value="IRR").font = Font(bold=True)
        ws.cell(row=returns_row+2, column=2, value=f"{preview['irr']:.1%}")
        
        ws.cell(row=returns_row+3, column=1, value="MOIC").font = Font(bold=True)
        ws.cell(row=returns_row+3, column=2, value=f"{preview['moic']:.1f}x")
        
        ws.cell(row=returns_row+4, column=1, value="Entry EV").font = Font(bold=True)
        ws.cell(row=returns_row+4, column=2, value=preview["entry_ev"])
        
        ws.cell(row=returns_row+5, column=1, value="Exit Value").font = Font(bold=True)
        ws.cell(row=returns_row+5, column=2, value=preview["exit_value"])
        
        # Auto-fit columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_comps_sheet(self, wb: Workbook, assumptions_result, preview: Dict[str, Any]):
        """Create Trading Comps sheet"""
        ws = wb.create_sheet("Trading Comps")
        
        # Peer table
        ws.cell(row=1, column=1, value="Trading Comparables").font = Font(size=14, bold=True)
        
        headers = ["Company", "Ticker", "EV/Revenue", "EV/EBITDA", "P/E"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        peers = preview["peers"]
        for row, peer in enumerate(peers, 4):
            ws.cell(row=row, column=1, value=peer["name"]).border = self.thin_border
            ws.cell(row=row, column=2, value=peer["ticker"]).border = self.thin_border
            ws.cell(row=row, column=3, value=peer["ev_revenue"]).border = self.thin_border
            ws.cell(row=row, column=4, value=peer["ev_ebitda"]).border = self.thin_border
            ws.cell(row=row, column=5, value=peer["pe"]).border = self.thin_border
        
        # Medians
        median_row = len(peers) + 6
        ws.cell(row=median_row, column=1, value="MEDIAN").font = Font(bold=True)
        ws.cell(row=median_row, column=3, value=preview["median"]["ev_revenue"]).border = self.thin_border
        ws.cell(row=median_row, column=4, value=preview["median"]["ev_ebitda"]).border = self.thin_border
        ws.cell(row=median_row, column=5, value=preview["median"]["pe"]).border = self.thin_border
        
        # Implied valuation
        implied_row = median_row + 3
        ws.cell(row=implied_row, column=1, value="Implied Valuation Range").font = Font(size=14, bold=True)
        
        implied_range = preview["implied_range"]
        ws.cell(row=implied_row+2, column=1, value="Low").font = Font(bold=True)
        ws.cell(row=implied_row+2, column=2, value=implied_range["low"]).border = self.thin_border
        
        ws.cell(row=implied_row+3, column=1, value="Mid").font = Font(bold=True)
        ws.cell(row=implied_row+3, column=2, value=implied_range["mid"]).border = self.thin_border
        
        ws.cell(row=implied_row+4, column=1, value="High").font = Font(bold=True)
        ws.cell(row=implied_row+4, column=2, value=implied_range["high"]).border = self.thin_border
        
        # Auto-fit columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_merger_sheet(self, wb: Workbook, assumptions_result, preview: Dict[str, Any]):
        """Create Merger model sheet"""
        ws = wb.create_sheet("Merger Model")
        
        # Pro forma income statement
        ws.cell(row=1, column=1, value="Pro Forma Income Statement").font = Font(size=14, bold=True)
        
        headers = ["Metric", "Acquirer", "Target", "Synergies", "Pro Forma"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        # Pro forma metrics
        acquirer_metrics = preview["acquirer_metrics"]
        target_metrics = preview["target_metrics"]
        synergies = preview["synergies"]
        pf_metrics = preview["pf_metrics"]
        
        metrics = [
            ("Revenue", acquirer_metrics["revenue"], target_metrics["revenue"], synergies["revenue_synergies"], pf_metrics["revenue"]),
            ("EBITDA", acquirer_metrics["ebitda"], target_metrics["ebitda"], synergies["total_synergies"], pf_metrics["ebitda"]),
            ("Net Income", acquirer_metrics["net_income"], target_metrics["net_income"], synergies["total_synergies"] * 0.7, pf_metrics["net_income"]),
            ("Shares", acquirer_metrics["shares"], 0, 0, pf_metrics["shares"]),
            ("EPS", acquirer_metrics["eps"], 0, 0, pf_metrics["eps"])
        ]
        
        for row, (metric, acquirer, target, synergy, pf) in enumerate(metrics, 4):
            ws.cell(row=row, column=1, value=metric).border = self.thin_border
            ws.cell(row=row, column=2, value=acquirer).border = self.thin_border
            ws.cell(row=row, column=3, value=target).border = self.thin_border
            ws.cell(row=row, column=4, value=synergy).border = self.thin_border
            ws.cell(row=row, column=5, value=pf).border = self.thin_border
        
        # Financing
        financing_row = len(metrics) + 6
        ws.cell(row=financing_row, column=1, value="Financing").font = Font(size=14, bold=True)
        
        financing = preview["financing"]
        ws.cell(row=financing_row+2, column=1, value="Purchase Price").font = Font(bold=True)
        ws.cell(row=financing_row+2, column=2, value=financing["purchase_price"]).border = self.thin_border
        
        ws.cell(row=financing_row+3, column=1, value="Debt Financing").font = Font(bold=True)
        ws.cell(row=financing_row+3, column=2, value=financing["debt_financing"]).border = self.thin_border
        
        ws.cell(row=financing_row+4, column=1, value="Equity Financing").font = Font(bold=True)
        ws.cell(row=financing_row+4, column=2, value=financing["equity_financing"]).border = self.thin_border
        
        # Accretion/Dilution
        accretion_row = financing_row + 6
        ws.cell(row=accretion_row, column=1, value="EPS Accretion/Dilution").font = Font(size=14, bold=True)
        
        eps_accretion = preview["eps_accretion"]
        ws.cell(row=accretion_row+2, column=1, value="EPS Accretion %").font = Font(bold=True)
        cell = ws.cell(row=accretion_row+2, column=2, value=f"{eps_accretion:.1f}%")
        cell.border = self.thin_border
        
        if eps_accretion > 0:
            cell.font = Font(color="00FF00", bold=True)  # Green for accretion
        else:
            cell.font = Font(color="FF0000", bold=True)  # Red for dilution
        
        # Auto-fit columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_summary_sheet(self, wb: Workbook, assumptions_result, preview: Dict[str, Any], model_type: str):
        """Create summary sheet with checks"""
        ws = wb.create_sheet("Summary")
        
        # Model summary
        ws.cell(row=1, column=1, value=f"{model_type} Model Summary").font = Font(size=16, bold=True)
        ws.cell(row=1, column=2, value=assumptions_result.ticker).font = Font(size=16, bold=True)
        ws.cell(row=1, column=3, value=datetime.now().strftime("%Y-%m-%d")).font = Font(size=16, bold=True)
        
        # Key metrics based on model type
        if model_type == "DCF":
            ws.cell(row=3, column=1, value="Enterprise Value").font = Font(bold=True)
            ws.cell(row=3, column=2, value=preview["ev"]).border = self.thin_border
            
            ws.cell(row=4, column=1, value="Equity Value").font = Font(bold=True)
            ws.cell(row=4, column=2, value=preview["equity_value"]).border = self.thin_border
            
            ws.cell(row=5, column=1, value="Implied Price").font = Font(bold=True)
            ws.cell(row=5, column=2, value=preview["implied_price"]).border = self.thin_border
            
            ws.cell(row=6, column=1, value="Upside %").font = Font(bold=True)
            cell = ws.cell(row=6, column=2, value=f"{preview['upside_pct']:.1f}%")
            cell.border = self.thin_border
            if preview["upside_pct"] > 0:
                cell.font = Font(color="00FF00", bold=True)
            else:
                cell.font = Font(color="FF0000", bold=True)
        
        elif model_type == "LBO":
            ws.cell(row=3, column=1, value="IRR").font = Font(bold=True)
            ws.cell(row=3, column=2, value=f"{preview['irr']:.1%}").border = self.thin_border
            
            ws.cell(row=4, column=1, value="MOIC").font = Font(bold=True)
            ws.cell(row=4, column=2, value=f"{preview['moic']:.1f}x").border = self.thin_border
            
            ws.cell(row=5, column=1, value="Entry EV").font = Font(bold=True)
            ws.cell(row=5, column=2, value=preview["entry_ev"]).border = self.thin_border
            
            ws.cell(row=6, column=1, value="Exit Value").font = Font(bold=True)
            ws.cell(row=6, column=2, value=preview["exit_value"]).border = self.thin_border
        
        # Checks section
        checks_row = 10
        ws.cell(row=checks_row, column=1, value="Model Checks").font = Font(size=14, bold=True)
        
        checks = []
        if model_type == "DCF":
            checks = [
                ("Terminal Growth < WACC", preview["terminal_growth"] < preview["wacc"]),
                ("TV Share < 85%", preview["pv_terminal"] / preview["ev"] < 0.85),
                ("WACC in Range", 0.06 <= preview["wacc"] <= 0.14)
            ]
        elif model_type == "LBO":
            checks = [
                ("Sources = Uses", abs(sum(preview["sources_uses"]["sources"].values()) - sum(preview["sources_uses"]["uses"].values())) < 1),
                ("IRR > 15%", preview["irr"] > 0.15),
                ("MOIC > 2.0x", preview["moic"] > 2.0)
            ]
        
        for row, (check_name, passed) in enumerate(checks, checks_row + 2):
            ws.cell(row=row, column=1, value=check_name).border = self.thin_border
            cell = ws.cell(row=row, column=2, value="✓" if passed else "✗")
            cell.border = self.thin_border
            if passed:
                cell.font = Font(color="00FF00", bold=True)
            else:
                cell.font = Font(color="FF0000", bold=True)
        
        # Auto-fit columns
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 20
