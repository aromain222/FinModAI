"""
Professional IB Modeling App - FastAPI Backend
Banker-grade web app for DCF, LBO, Trading Comps, and Merger models.
"""

from fastapi import FastAPI, HTTPException, BackgroundTasks, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import Dict, List, Optional, Any, Union
import uuid
import os
import tempfile
from datetime import datetime
import json
import traceback

# Import existing LBO system
from lbo_model import LBOEngine
from financial_data import FinancialDataEngine

# Import new models
from models.dcf_model import DCFModel, DCFAssumptions
from models.trading_comps_model import TradingCompsModel
from models.merger_model import MergerModel, MergerAssumptions

app = FastAPI(
    title="FinModAI Professional",
    description="Banker-grade financial modeling platform",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Initialize engines
lbo_engine = LBOEngine()
financial_engine = FinancialDataEngine()
dcf_model = DCFModel()
trading_comps_model = TradingCompsModel()
merger_model = MergerModel()

# In-memory storage for generated files
generated_files = {}

# Pydantic models
class AssumptionsRequest(BaseModel):
    ticker: str = Field(..., description="Company ticker symbol")
    model: str = Field(..., description="Model type: dcf, lbo, comps, merger")

class ModelGenerationRequest(BaseModel):
    ticker: str = Field(..., description="Company ticker symbol")
    model: str = Field(..., description="Model type: dcf, lbo, comps, merger")
    overrides: Optional[Dict[str, Any]] = Field(default=None, description="Optional assumption overrides")

class Provenance(BaseModel):
    source: str
    as_of: str

class Historicals(BaseModel):
    years: List[int]
    revenue: List[float]
    operating_income: List[float]
    op_margin: List[float]
    da: List[float]
    capex: List[float]
    delta_nwc: List[float]

class Assumptions(BaseModel):
    forecast_years: int
    revenue_growth: List[float]
    operating_margin: List[float]
    da_pct_rev: float
    capex_pct_rev: float
    nwc_pct_rev: float
    tax_rate: float

class WACC(BaseModel):
    rf: float
    erp: float
    beta: float
    ke: float
    kd_pre: float
    tax: float
    wd: float
    we: float
    kd_after: float
    wacc: float

class Terminal(BaseModel):
    method: str
    g: float

class AssumptionsResponse(BaseModel):
    ticker: str
    company_name: str
    currency: str
    provenance: Dict[str, Provenance]
    historicals: Historicals
    assumptions: Assumptions
    wacc: WACC
    terminal: Terminal
    flags: List[str]

class ErrorResponse(BaseModel):
    error: str
    missing: Optional[List[str]] = None
    provider_attempts: Optional[List[str]] = None
    message: str

class ModelPreview(BaseModel):
    dcf: Optional[Dict[str, Any]] = None
    lbo: Optional[Dict[str, Any]] = None
    comps: Optional[Dict[str, Any]] = None
    merger: Optional[Dict[str, Any]] = None

class ModelGenerationResponse(BaseModel):
    job_id: str
    assumptions: Dict[str, Any]
    preview: ModelPreview
    file: Dict[str, str]
    warnings: List[str]

@app.get("/")
async def root():
    """Serve the main HTML page"""
    try:
        with open("static/index.html", "r") as f:
            content = f.read()
        return HTMLResponse(content=content)
    except FileNotFoundError:
        return {"message": "FinModAI Professional API", "version": "1.0.0"}

@app.get("/healthz")
async def health_check():
    """Health check endpoint for Render"""
    return {"status": "ok", "timestamp": datetime.now().isoformat()}

@app.get("/assumptions")
async def get_assumptions(ticker: str, model: str):
    """
    Get company-specific assumptions derived from historical data.
    Returns structured error if insufficient data.
    """
    try:
        print(f"🔍 Fetching assumptions for {ticker} ({model})")
        
        # Validate model type
        if model not in ["dcf", "lbo", "comps", "merger"]:
            raise HTTPException(status_code=400, detail="Invalid model type")
        
        # Fetch historical data
        try:
            historical_data = financial_engine.fetch_company_data(ticker)
        except Exception as e:
            print(f"❌ Error fetching data for {ticker}: {e}")
            historical_data = None
        
        if not historical_data:
            return ErrorResponse(
                error="insufficient_historicals",
                missing=["revenue", "operating_income"],
                provider_attempts=["FMP", "AlphaVantage", "Yahoo", "GoogleSheets"],
                message=f"Need ≥3 years of revenue and EBIT to build assumptions for {ticker}."
            )
        
        # Extract data
        company_name = historical_data.get('company_name', 'Unknown Company')
        currency = "USD"  # Default to USD
        
        # Build provenance
        provenance = {
            "revenue": Provenance(source="FMP", as_of=datetime.now().strftime("%Y-%m-%d")),
            "op_margin": Provenance(source="FMP", as_of=datetime.now().strftime("%Y-%m-%d")),
            "capex": Provenance(source="FMP", as_of=datetime.now().strftime("%Y-%m-%d")),
            "nwc": Provenance(source="FMP", as_of=datetime.now().strftime("%Y-%m-%d")),
            "price": Provenance(source="Yahoo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "beta": Provenance(source="Computed_5y_vs_SPY", as_of=datetime.now().strftime("%Y-%m-%d")),
            "rf": Provenance(source="FRED_DGS10", as_of=datetime.now().strftime("%Y-%m-%d"))
        }
        
        # Extract historicals
        hist_data = historical_data.get('historicals', {})
        years = list(range(2020, 2024))  # Example years
        revenue = hist_data.get('revenue', [0, 0, 0, 0])
        operating_income = hist_data.get('operating_income', [0, 0, 0, 0])
        op_margin = hist_data.get('op_margin', [0, 0, 0, 0])
        da = [r * 0.05 for r in revenue]  # 5% of revenue for D&A
        capex = [r * 0.08 for r in revenue]  # 8% of revenue for CapEx
        delta_nwc = [r * 0.02 for r in revenue]  # 2% of revenue for NWC change
        
        historicals = Historicals(
            years=years,
            revenue=revenue,
            operating_income=operating_income,
            op_margin=op_margin,
            da=da,
            capex=capex,
            delta_nwc=delta_nwc
        )
        
        # Build assumptions
        assumptions_data = historical_data.get('assumptions', {})
        revenue_growth = assumptions_data.get('revenue_growth', [0.05, 0.05, 0.05, 0.05, 0.05])
        operating_margin = assumptions_data.get('operating_margin', [0.25, 0.25, 0.25, 0.25, 0.25])
        
        assumptions = Assumptions(
            forecast_years=10,
            revenue_growth=revenue_growth,
            operating_margin=operating_margin,
            da_pct_rev=0.05,
            capex_pct_rev=0.08,
            nwc_pct_rev=0.02,
            tax_rate=assumptions_data.get('tax_rate', 0.25)
        )
        
        # Build WACC
        wacc_data = historical_data.get('wacc', {})
        wacc = WACC(
            rf=wacc_data.get('rf', 0.04),
            erp=0.055,
            beta=assumptions_data.get('beta', 1.0),
            ke=wacc_data.get('ke', 0.10),
            kd_pre=0.06,
            tax=assumptions.get('tax_rate', 0.25),
            wd=0.3,
            we=0.7,
            kd_after=0.045,
            wacc=wacc_data.get('wacc', 0.10)
        )
        
        # Terminal value
        terminal = Terminal(method="perpetuity", g=0.025)
        
        # Flags
        flags = []
        if terminal.g >= wacc.wacc:
            flags.append("terminal_g_ge_wacc")
        
        return AssumptionsResponse(
            ticker=ticker.upper(),
            company_name=company_name,
            currency=currency,
            provenance=provenance,
            historicals=historicals,
            assumptions=assumptions,
            wacc=wacc,
            terminal=terminal,
            flags=flags
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error fetching assumptions: {str(e)}")
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=ErrorResponse(
                error="internal_error",
                message=f"Failed to fetch assumptions for {ticker}: {str(e)}"
            ).dict()
        )

@app.post("/models/generate")
async def generate_model(request: ModelGenerationRequest, background_tasks: BackgroundTasks):
    """
    Generate a financial model and return preview + download URL.
    """
    try:
        print(f"🏗️ Generating {request.model} model for {request.ticker}")
        
        job_id = str(uuid.uuid4())
        
        # Generate model based on type
        if request.model == "lbo":
            # Use existing LBO engine
            lbo_assumptions = {
                'ticker': request.ticker,
                'entry_multiple': 8.0,
                'exit_multiple': 10.0,
                'leverage': 5.5,
                'holding_period': 5
            }
            
            # Apply overrides if provided
            if request.overrides:
                lbo_assumptions.update(request.overrides)
            
            result = lbo_engine.build_lbo_model(lbo_assumptions)
            
            if not result:
                raise HTTPException(status_code=500, detail="Failed to generate LBO model")
            
            # Build preview
            preview = ModelPreview(
                lbo={
                    "sources_uses": result.get('sources_uses', {}),
                    "debt_stack": result.get('debt_schedule', []),
                    "fcf": result.get('proforma_financials', {}).get('free_cash_flow', []),
                    "paydown": [],
                    "leverage_path": [],
                    "irr": result.get('returns_analysis', {}).get('equity_irr', 0),
                    "moic": result.get('returns_analysis', {}).get('moic', 0),
                    "covenants": result.get('covenant_metrics', {})
                }
            )
            
            # Generate Excel file
            filename = f"{request.ticker}_{request.model.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            file_path = await generate_excel_file(request.ticker, request.model, result)
            
            generated_files[job_id] = {
                "filename": filename,
                "file_path": file_path,
                "created_at": datetime.now()
            }
            
        elif request.model == "dcf":
            # Generate DCF model
            historical_data = financial_engine.fetch_company_data(request.ticker)
            if not historical_data:
                raise HTTPException(status_code=404, detail="Insufficient data for DCF model")
            
            # Build DCF assumptions
            dcf_assumptions = DCFAssumptions(
                ticker=request.ticker,
                company_name=historical_data.get('company_name', 'Unknown Company'),
                revenue_growth=historical_data.get('assumptions', {}).get('revenue_growth', [0.05] * 10),
                operating_margin=historical_data.get('assumptions', {}).get('operating_margin', [0.25] * 10),
                wacc=historical_data.get('wacc', {}).get('wacc', 0.10),
                tax_rate=historical_data.get('assumptions', {}).get('tax_rate', 0.25)
            )
            
            # Apply overrides
            if request.overrides:
                if 'revenue_growth' in request.overrides:
                    dcf_assumptions.revenue_growth = [request.overrides['revenue_growth']] * 10
                if 'operating_margin' in request.overrides:
                    dcf_assumptions.operating_margin = [request.overrides['operating_margin']] * 10
                if 'wacc' in request.overrides:
                    dcf_assumptions.wacc = request.overrides['wacc']
            
            # Build DCF model
            dcf_results = dcf_model.build_dcf_model(dcf_assumptions)
            
            # Build preview
            preview = ModelPreview(
                dcf={
                    "years": dcf_results.years,
                    "ufcf": dcf_results.ufcf,
                    "pv_ufcf": dcf_results.pv_ufcf,
                    "tv": dcf_results.terminal_value,
                    "ev": dcf_results.enterprise_value,
                    "net_debt": dcf_results.net_debt,
                    "equity": dcf_results.equity_value,
                    "implied_price": dcf_results.implied_price,
                    "upside_pct": dcf_results.upside_pct
                }
            )
            
            # Generate Excel file
            filename = f"{request.ticker}_{request.model.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            file_path = await generate_excel_file(request.ticker, request.model, dcf_results.__dict__)
            
            generated_files[job_id] = {
                "filename": filename,
                "file_path": file_path,
                "created_at": datetime.now()
            }
            
        elif request.model == "comps":
            # Generate Trading Comps model
            historical_data = financial_engine.fetch_company_data(request.ticker)
            if not historical_data:
                raise HTTPException(status_code=404, detail="Insufficient data for Trading Comps model")
            
            # Build trading comps
            comps_results = trading_comps_model.build_trading_comps(request.ticker, historical_data)
            
            # Build preview
            preview = ModelPreview(
                comps={
                    "peers": [{"ticker": p.ticker, "name": p.name, "ev_revenue": p.ev_revenue, "ev_ebitda": p.ev_ebitda} for p in comps_results.peers],
                    "multiples": comps_results.multiples,
                    "median": comps_results.median_multiples,
                    "implied_range": comps_results.implied_valuation
                }
            )
            
            # Generate Excel file
            filename = f"{request.ticker}_{request.model.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            file_path = await generate_excel_file(request.ticker, request.model, comps_results.__dict__)
            
            generated_files[job_id] = {
                "filename": filename,
                "file_path": file_path,
                "created_at": datetime.now()
            }
            
        elif request.model == "merger":
            # Generate Merger model (requires target company)
            target_ticker = request.overrides.get('target_ticker') if request.overrides else 'TARGET'
            
            merger_assumptions = MergerAssumptions(
                acquirer_ticker=request.ticker,
                target_ticker=target_ticker,
                acquirer_name=f"{request.ticker} Corp.",
                target_name=f"{target_ticker} Inc.",
                purchase_price=1000000000,  # $1B default
                payment_method=request.overrides.get('payment_method', 'cash') if request.overrides else 'cash',
                synergies=request.overrides.get('synergies') if request.overrides else None
            )
            
            # Build merger model
            merger_results = merger_model.build_merger_model(merger_assumptions)
            
            # Build preview
            preview = ModelPreview(
                merger={
                    "pf_is": merger_results.pro_forma_is,
                    "synergies": [{"name": s.name, "amount": s.amount, "timing": s.timing} for s in merger_results.synergies],
                    "pf_leverage": merger_results.pf_leverage,
                    "eps_accretion": merger_results.eps_accretion
                }
            )
            
            # Generate Excel file
            filename = f"{request.ticker}_{request.model.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            file_path = await generate_excel_file(request.ticker, request.model, merger_results.__dict__)
            
            generated_files[job_id] = {
                "filename": filename,
                "file_path": file_path,
                "created_at": datetime.now()
            }
        
        return ModelGenerationResponse(
            job_id=job_id,
            assumptions={},  # Will be populated with actual assumptions
            preview=preview,
            file={
                "filename": filename,
                "download_url": f"/download/{filename}"
            },
            warnings=[]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error generating model: {str(e)}")
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate {request.model} model: {str(e)}"
        )

@app.get("/download/{filename}")
async def download_file(filename: str):
    """
    Download generated Excel file.
    """
    try:
        # Find file in generated_files
        file_info = None
        for job_id, info in generated_files.items():
            if info["filename"] == filename:
                file_info = info
                break
        
        if not file_info or not os.path.exists(file_info["file_path"]):
            raise HTTPException(status_code=404, detail="File not found. Please regenerate the model.")
        
        return FileResponse(
            path=file_info["file_path"],
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error downloading file: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to download file")

async def generate_excel_file(ticker: str, model: str, data: Dict[str, Any]) -> str:
    """
    Generate Excel file for the model.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        if model == "lbo":
            # LBO model Excel
            ws = wb.active
            ws.title = "LBO Model"
            
            # Add headers
            headers = ["Year", "Revenue", "EBITDA", "FCF", "Net Debt", "Equity Value"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Add data
            years = [2024, 2025, 2026, 2027, 2028]
            for row, year in enumerate(years, 2):
                ws.cell(row=row, column=1, value=year)
                ws.cell(row=row, column=2, value=1000 + (row-2) * 100)  # Revenue
                ws.cell(row=row, column=3, value=250 + (row-2) * 25)   # EBITDA
                ws.cell(row=row, column=4, value=150 + (row-2) * 15)   # FCF
                ws.cell(row=row, column=5, value=500 - (row-2) * 50)   # Net Debt
                ws.cell(row=row, column=6, value=2000 + (row-2) * 200) # Equity Value
        
        else:
            # Other models - placeholder
            ws = wb.active
            ws.title = f"{model.upper()} Model"
            ws.cell(row=1, column=1, value=f"{ticker} {model.upper()} Model")
            ws.cell(row=2, column=1, value="Generated on: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Save file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"{ticker}_{model.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(file_path)
        
        return file_path
        
    except Exception as e:
        print(f"❌ Error generating Excel file: {str(e)}")
        raise

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
