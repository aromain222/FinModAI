#!/usr/bin/env python3
"""
Verification script to ensure all dependencies are properly installed
before deployment. Run this locally to catch issues early.
"""

import sys

def verify_imports():
    """Verify all critical imports work."""
    print("="*70)
    print("DEPLOYMENT VERIFICATION - Testing Critical Dependencies")
    print("="*70)
    print()
    
    critical_modules = [
        ('flask', 'Flask web framework'),
        ('gunicorn', 'Gunicorn WSGI server'),
        ('pandas', 'Pandas data analysis'),
        ('numpy', 'NumPy numerical computing'),
        ('yfinance', 'Yahoo Finance data'),
        ('requests', 'HTTP requests'),
        ('bs4', 'BeautifulSoup HTML parsing'),
        ('et_xmlfile', 'ET XML file support (openpyxl dependency)'),
        ('openpyxl', 'Excel file support'),
    ]
    
    passed = []
    failed = []
    
    for module_name, description in critical_modules:
        try:
            module = __import__(module_name)
            version = getattr(module, '__version__', 'unknown')
            print(f"✓ {module_name:20s} v{version:15s} - {description}")
            passed.append(module_name)
        except ImportError as e:
            print(f"✗ {module_name:20s} FAILED - {description}")
            print(f"  Error: {e}")
            failed.append(module_name)
    
    print()
    print("="*70)
    
    if failed:
        print(f"❌ VERIFICATION FAILED - {len(failed)} module(s) missing:")
        for module in failed:
            print(f"   - {module}")
        print()
        print("To fix, run:")
        print("   pip install -r requirements.txt")
        print()
        return False
    else:
        print(f"✅ ALL CHECKS PASSED - {len(passed)} modules verified")
        print()
        print("Safe to deploy!")
        print()
        return True

def verify_openpyxl_functionality():
    """Test that openpyxl actually works."""
    print("="*70)
    print("Testing openpyxl functionality...")
    print("="*70)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Add some data with formatting
        ws['A1'] = "Test"
        ws['A1'].font = Font(bold=True)
        ws['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A1'].border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        print("✓ Created workbook")
        print("✓ Applied styling")
        print("✓ openpyxl is fully functional")
        print()
        return True
        
    except Exception as e:
        print(f"✗ openpyxl functionality test failed: {e}")
        print()
        return False

def verify_minimal_app():
    """Test that minimal_app.py can be imported."""
    print("="*70)
    print("Testing minimal_app.py import...")
    print("="*70)
    
    try:
        import minimal_app
        print("✓ minimal_app.py imported successfully")
        print(f"✓ Flask app object: {minimal_app.app}")
        print()
        return True
    except Exception as e:
        print(f"✗ Failed to import minimal_app.py: {e}")
        import traceback
        traceback.print_exc()
        print()
        return False

def main():
    """Run all verification checks."""
    results = []
    
    # Run checks
    results.append(("Import Verification", verify_imports()))
    results.append(("openpyxl Functionality", verify_openpyxl_functionality()))
    results.append(("minimal_app Import", verify_minimal_app()))
    
    # Summary
    print()
    print("="*70)
    print("VERIFICATION SUMMARY")
    print("="*70)
    
    all_passed = True
    for check_name, passed in results:
        status = "✅ PASS" if passed else "❌ FAIL"
        print(f"{status} - {check_name}")
        if not passed:
            all_passed = False
    
    print("="*70)
    
    if all_passed:
        print()
        print("🎉 ALL VERIFICATIONS PASSED!")
        print("Your application is ready for deployment.")
        print()
        sys.exit(0)
    else:
        print()
        print("⚠️  VERIFICATION FAILED!")
        print("Please fix the issues above before deploying.")
        print()
        sys.exit(1)

if __name__ == '__main__':
    main()

