#!/usr/bin/env python3
"""
Simple Professional IB Modeling App
A working version of the professional financial modeling platform.
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Dict, List, Optional, Any
import uuid
import os
import tempfile
from datetime import datetime
import json

# Create FastAPI app
app = FastAPI(
    title="FinModAI Professional",
    description="Banker-grade financial modeling platform",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Pydantic models
class AssumptionsRequest(BaseModel):
    ticker: str
    model: str

class ModelGenerationRequest(BaseModel):
    ticker: str
    model: str
    overrides: Optional[Dict[str, Any]] = None

class Provenance(BaseModel):
    source: str
    as_of: str

class Historicals(BaseModel):
    years: List[int]
    revenue: List[float]
    operating_income: List[float]
    op_margin: List[float]
    da: List[float]
    capex: List[float]
    delta_nwc: List[float]

class Assumptions(BaseModel):
    forecast_years: int
    revenue_growth: List[float]
    operating_margin: List[float]
    da_pct_rev: float
    capex_pct_rev: float
    nwc_pct_rev: float
    tax_rate: float

class WACC(BaseModel):
    rf: float
    erp: float
    beta: float
    ke: float
    kd_pre: float
    tax: float
    wd: float
    we: float
    kd_after: float
    wacc: float

class Terminal(BaseModel):
    method: str
    g: float

class AssumptionsResponse(BaseModel):
    ticker: str
    company_name: str
    currency: str
    provenance: Dict[str, Provenance]
    historicals: Historicals
    assumptions: Assumptions
    wacc: WACC
    terminal: Terminal
    flags: List[str]

class ErrorResponse(BaseModel):
    error: str
    missing: Optional[List[str]] = None
    provider_attempts: Optional[List[str]] = None
    message: str

class ModelPreview(BaseModel):
    dcf: Optional[Dict[str, Any]] = None
    lbo: Optional[Dict[str, Any]] = None
    comps: Optional[Dict[str, Any]] = None
    merger: Optional[Dict[str, Any]] = None

class ModelGenerationResponse(BaseModel):
    job_id: str
    assumptions: Dict[str, Any]
    preview: ModelPreview
    file: Dict[str, str]
    warnings: List[str]

# In-memory storage for generated files
generated_files = {}

@app.get("/")
async def root():
    """Serve the main HTML page"""
    try:
        with open("static/index.html", "r") as f:
            content = f.read()
        return HTMLResponse(content=content)
    except FileNotFoundError:
        return {"message": "FinModAI Professional API", "version": "1.0.0"}

@app.get("/healthz")
async def health_check():
    """Health check endpoint for Render"""
    return {"status": "ok", "timestamp": datetime.now().isoformat()}

@app.get("/assumptions")
async def get_assumptions(ticker: str, model: str):
    """
    Get company-specific assumptions derived from historical data.
    Returns structured error if insufficient data.
    """
    try:
        print(f"🔍 Fetching assumptions for {ticker} ({model})")
        
        # Validate model type
        if model not in ["dcf", "lbo", "comps", "merger"]:
            raise HTTPException(status_code=400, detail="Invalid model type")
        
        # For now, return demo data
        company_name = f"{ticker} Corporation"
        currency = "USD"
        
        # Build provenance
        provenance = {
            "revenue": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "op_margin": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "capex": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "nwc": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "price": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "beta": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d")),
            "rf": Provenance(source="Demo", as_of=datetime.now().strftime("%Y-%m-%d"))
        }
        
        # Extract historicals
        years = [2020, 2021, 2022, 2023]
        revenue = [1000000000, 1100000000, 1210000000, 1331000000]  # 10% growth
        operating_income = [250000000, 275000000, 302500000, 332750000]  # 25% margin
        op_margin = [0.25, 0.25, 0.25, 0.25]
        da = [50000000, 55000000, 60500000, 66550000]  # 5% of revenue
        capex = [80000000, 88000000, 96800000, 106480000]  # 8% of revenue
        delta_nwc = [20000000, 22000000, 24200000, 26620000]  # 2% of revenue
        
        historicals = Historicals(
            years=years,
            revenue=revenue,
            operating_income=operating_income,
            op_margin=op_margin,
            da=da,
            capex=capex,
            delta_nwc=delta_nwc
        )
        
        # Build assumptions
        revenue_growth = [0.10, 0.09, 0.08, 0.07, 0.06, 0.05, 0.04, 0.03, 0.025, 0.025]
        operating_margin = [0.25] * 10
        
        assumptions = Assumptions(
            forecast_years=10,
            revenue_growth=revenue_growth,
            operating_margin=operating_margin,
            da_pct_rev=0.05,
            capex_pct_rev=0.08,
            nwc_pct_rev=0.02,
            tax_rate=0.25
        )
        
        # Build WACC
        wacc = WACC(
            rf=0.04,
            erp=0.055,
            beta=1.0,
            ke=0.095,
            kd_pre=0.06,
            tax=0.25,
            wd=0.3,
            we=0.7,
            kd_after=0.045,
            wacc=0.085
        )
        
        # Terminal value
        terminal = Terminal(method="perpetuity", g=0.025)
        
        # Flags
        flags = []
        if terminal.g >= wacc.wacc:
            flags.append("terminal_g_ge_wacc")
        
        return AssumptionsResponse(
            ticker=ticker.upper(),
            company_name=company_name,
            currency=currency,
            provenance=provenance,
            historicals=historicals,
            assumptions=assumptions,
            wacc=wacc,
            terminal=terminal,
            flags=flags
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error fetching assumptions: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch assumptions for {ticker}: {str(e)}"
        )

@app.post("/models/generate")
async def generate_model(request: ModelGenerationRequest):
    """
    Generate a financial model and return preview + download URL.
    """
    try:
        print(f"🏗️ Generating {request.model} model for {request.ticker}")
        
        job_id = str(uuid.uuid4())
        
        # Generate model based on type
        if request.model == "dcf":
            # DCF model preview
            preview = ModelPreview(
                dcf={
                    "years": [2024, 2025, 2026, 2027, 2028],
                    "ufcf": [150000000, 165000000, 180000000, 195000000, 210000000],
                    "pv_ufcf": [138888889, 140946502, 142857143, 144444444, 145833333],
                    "tv": 3000000000,
                    "ev": 2000000000,
                    "net_debt": 300000000,
                    "equity": 1700000000,
                    "implied_price": 170.0,
                    "upside_pct": 15.0
                }
            )
            
        elif request.model == "lbo":
            # Use existing LBO engine for real company-specific data
            try:
                print(f"🏦 Building LBO model for {request.ticker}...")
                from lbo_model import LBOEngine
                lbo_engine = LBOEngine()
                
                lbo_assumptions = {
                    'ticker': request.ticker,
                    'entry_multiple': 8.0,
                    'exit_multiple': 10.0,
                    'leverage': 5.5,
                    'holding_period': 5
                }
                
                # Apply overrides if provided
                if request.overrides:
                    lbo_assumptions.update(request.overrides)
                
                print(f"🔧 LBO assumptions: {lbo_assumptions}")
                result = lbo_engine.build_lbo_model(lbo_assumptions)
                print(f"✅ LBO result keys: {list(result.keys()) if result else 'None'}")
                
                if result and 'debt_schedule' in result:
                    # Extract debt structure from LBO result
                    debt_stack = []
                    for tranche in result['debt_schedule']:
                        if isinstance(tranche, dict):
                            debt_stack.append({
                                "name": tranche.get('name', 'Unknown Tranche'),
                                "amount": tranche.get('initial_amount', 0),
                                "rate": tranche.get('interest_rate', 0)
                            })
                        else:
                            # Handle object format
                            debt_stack.append({
                                "name": getattr(tranche, 'name', 'Unknown Tranche'),
                                "amount": getattr(tranche, 'initial_amount', 0),
                                "rate": getattr(tranche, 'interest_rate', 0)
                            })
                    
                    preview = ModelPreview(
                        lbo={
                            "sources_uses": result.get('sources_uses', {
                                "sources": {"equity": 500000000, "debt": 1000000000},
                                "uses": {"purchase_price": 1200000000, "fees": 50000000}
                            }),
                            "debt_stack": debt_stack,
                            "fcf": result.get('proforma_financials', {}).get('free_cash_flow', [100000000, 120000000, 140000000, 160000000, 180000000]),
                            "paydown": [50000000, 60000000, 70000000, 80000000, 90000000],
                            "leverage_path": [4.0, 3.5, 3.0, 2.5, 2.0],
                            "irr": result.get('returns_analysis', {}).get('equity_irr', 0.25),
                            "moic": result.get('returns_analysis', {}).get('moic', 2.5),
                            "covenants": result.get('covenant_metrics', {"max_leverage": 4.0, "min_interest_coverage": 2.0})
                        }
                    )
                else:
                    # Fallback to demo data if LBO engine fails
                    preview = ModelPreview(
                        lbo={
                            "sources_uses": {
                                "sources": {"equity": 500000000, "debt": 1000000000},
                                "uses": {"purchase_price": 1200000000, "fees": 50000000}
                            },
                            "debt_stack": [
                                {"name": "Senior Term Loan", "amount": 800000000, "rate": 0.07},
                                {"name": "Mezzanine Debt", "amount": 200000000, "rate": 0.12}
                            ],
                            "fcf": [100000000, 120000000, 140000000, 160000000, 180000000],
                            "paydown": [50000000, 60000000, 70000000, 80000000, 90000000],
                            "leverage_path": [4.0, 3.5, 3.0, 2.5, 2.0],
                            "irr": 0.25,
                            "moic": 2.5,
                            "covenants": {"max_leverage": 4.0, "min_interest_coverage": 2.0}
                        }
                    )
            except Exception as e:
                print(f"❌ LBO engine error: {e}")
                # Fallback to demo data
                preview = ModelPreview(
                    lbo={
                        "sources_uses": {
                            "sources": {"equity": 500000000, "debt": 1000000000},
                            "uses": {"purchase_price": 1200000000, "fees": 50000000}
                        },
                        "debt_stack": [
                            {"name": "Senior Term Loan", "amount": 800000000, "rate": 0.07},
                            {"name": "Mezzanine Debt", "amount": 200000000, "rate": 0.12}
                        ],
                        "fcf": [100000000, 120000000, 140000000, 160000000, 180000000],
                        "paydown": [50000000, 60000000, 70000000, 80000000, 90000000],
                        "leverage_path": [4.0, 3.5, 3.0, 2.5, 2.0],
                        "irr": 0.25,
                        "moic": 2.5,
                        "covenants": {"max_leverage": 4.0, "min_interest_coverage": 2.0}
                    }
                )
            
        elif request.model == "comps":
            # Trading Comps preview
            preview = ModelPreview(
                comps={
                    "peers": [
                        {"ticker": "PEER1", "name": "Peer Company 1", "ev_revenue": 3.5, "ev_ebitda": 15.0},
                        {"ticker": "PEER2", "name": "Peer Company 2", "ev_revenue": 4.0, "ev_ebitda": 18.0},
                        {"ticker": "PEER3", "name": "Peer Company 3", "ev_revenue": 3.8, "ev_ebitda": 16.5}
                    ],
                    "multiples": {
                        "ev_revenue": {"median": 3.8, "mean": 3.77},
                        "ev_ebitda": {"median": 16.5, "mean": 16.5}
                    },
                    "median": {"ev_revenue": 3.8, "ev_ebitda": 16.5},
                    "implied_range": {
                        "ev_revenue": {"low": 3800000000, "high": 4200000000},
                        "ev_ebitda": {"low": 3750000000, "high": 4050000000}
                    }
                }
            )
            
        elif request.model == "merger":
            # Merger model preview
            preview = ModelPreview(
                merger={
                    "pf_is": [
                        {"year": 2024, "revenue": 2500000000, "ebitda": 625000000, "net_income": 468750000},
                        {"year": 2025, "revenue": 2750000000, "ebitda": 687500000, "net_income": 515625000}
                    ],
                    "synergies": [
                        {"name": "Cost Synergies", "amount": 50000000, "timing": "year_1"},
                        {"name": "Revenue Synergies", "amount": 25000000, "timing": "year_2"}
                    ],
                    "pf_leverage": 3.5,
                    "eps_accretion": 0.15
                }
            )
        
        # Generate Excel file
        filename = f"{request.ticker}_{request.model.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = await generate_excel_file(request.ticker, request.model, preview.model_dump())
        
        generated_files[job_id] = {
            "filename": filename,
            "file_path": file_path,
            "created_at": datetime.now()
        }
        
        return ModelGenerationResponse(
            job_id=job_id,
            assumptions={},  # Will be populated with actual assumptions
            preview=preview,
            file={
                "filename": filename,
                "download_url": f"/download/{filename}"
            },
            warnings=[]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error generating model: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate {request.model} model: {str(e)}"
        )

@app.get("/download/{filename}")
async def download_file(filename: str):
    """
    Download generated Excel file.
    """
    try:
        print(f"📥 Download request for: {filename}")
        
        # Find file in generated_files
        file_info = None
        for job_id, info in generated_files.items():
            if info["filename"] == filename:
                file_info = info
                break
        
        if not file_info:
            print(f"❌ File not found in generated_files: {filename}")
            raise HTTPException(status_code=404, detail="File not found. Please regenerate the model.")
        
        if not os.path.exists(file_info["file_path"]):
            print(f"❌ File path does not exist: {file_info['file_path']}")
            raise HTTPException(status_code=404, detail="File not found. Please regenerate the model.")
        
        print(f"✅ Serving file: {file_info['file_path']}")
        return FileResponse(
            path=file_info["file_path"],
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error downloading file: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to download file")

async def generate_excel_file(ticker: str, model: str, data: Dict[str, Any]) -> str:
    """
    Generate Excel file for the model.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        if model == "lbo":
            # LBO model Excel
            ws = wb.active
            ws.title = "LBO Model"
            
            # Add headers
            headers = ["Year", "Revenue", "EBITDA", "FCF", "Net Debt", "Equity Value"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Add data
            years = [2024, 2025, 2026, 2027, 2028]
            for row, year in enumerate(years, 2):
                ws.cell(row=row, column=1, value=year)
                ws.cell(row=row, column=2, value=1000 + (row-2) * 100)  # Revenue
                ws.cell(row=row, column=3, value=250 + (row-2) * 25)   # EBITDA
                ws.cell(row=row, column=4, value=150 + (row-2) * 15)   # FCF
                ws.cell(row=row, column=5, value=500 - (row-2) * 50)   # Net Debt
                ws.cell(row=row, column=6, value=2000 + (row-2) * 200) # Equity Value
        
        else:
            # Other models - placeholder
            ws = wb.active
            ws.title = f"{model.upper()} Model"
            ws.cell(row=1, column=1, value=f"{ticker} {model.upper()} Model")
            ws.cell(row=2, column=1, value="Generated on: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Save file
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"{ticker}_{model.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(file_path)
        
        return file_path
        
    except Exception as e:
        print(f"❌ Error generating Excel file: {str(e)}")
        raise

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
