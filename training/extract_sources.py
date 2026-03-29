#!/usr/bin/env python3
"""Populate extraction stubs with deterministic content from PDFs, workbooks, and manual overrides."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from pypdf import PdfReader
import xlrd

from training.build_extraction_stubs import build_stub
from training.corpus_logic import filter_sources, load_registry, load_templates

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "training" / "extracted_sources"

DATE_PATTERNS = [
    re.compile(r"\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}\b"),
    re.compile(r"\b\d{1,2}-\d{1,2}-\d{2,4}\b"),
]

MANUAL_OVERRIDES: dict[str, dict[str, Any]] = {
    "xylem_evoqua_sec_methodology": {
        "title": "Xylem / Evoqua Merger Valuation Disclosures",
        "deal_name": "Xylem / Evoqua merger",
        "filing_type": "8-K supplemental merger valuation disclosure",
        "advisors": ["Goldman Sachs & Co. LLC", "BofA Securities, Inc."],
        "valuation_analyses": [
            "Illustrative Discounted Cash Flow Analysis - Evoqua Standalone",
            "Illustrative Discounted Cash Flow Analysis - Pro Forma Combined Company",
            "Illustrative Present Value of Future Share Price Analysis",
            "Selected Transactions Analysis",
            "Selected Publicly Traded Companies Analysis",
        ],
        "peer_set": [
            "Ecolab Inc.",
            "Xylem Inc.",
            "Industrie de Nora S.p.A",
            "Pentair plc",
            "A.O. Smith Corporation",
            "Kurita Water Industries Ltd.",
            "Evoqua Water Technologies Corp.",
            "Franklin Electric Co., Inc.",
            "Badger Meter Inc",
            "Watts Water Technologies, Inc.",
            "Zurn Elkay Water Solutions Corporation",
            "Mueller Water Products, Inc.",
        ],
        "transaction_set": [
            "Zurn Water Solutions / Elkay Manufacturing",
            "Cott / Primo Water",
            "Culligan / AquaVenture",
            "Xylem / Pure Technologies",
            "SUEZ | CDPQ / GE Water & Power Technologies",
            "Xylem / Sensus USA",
            "Danaher / Pall",
            "Watts Water Technologies / AERCO International",
        ],
        "discount_rate_range": {
            "Goldman Sachs": "9.0% to 10.5%",
            "BofA Securities": "9.0% to 11.0%",
        },
        "terminal_multiple_range": {
            "Evoqua standalone": "18.0x to 22.0x",
            "Pro forma combined company": "21.0x to 25.0x",
        },
        "share_count": {
            "Evoqua fully diluted shares": "approximately 126.4 million to 126.8 million",
            "Pro forma combined fully diluted shares": "approximately 243.0 million to 243.2 million",
        },
        "net_debt": {
            "Evoqua standalone": "approximately $767 million",
            "Pro forma combined company": "approximately $1,742 million",
        },
        "implied_value_range": {
            "Evoqua standalone": "$42 to $54 per share",
        },
    }
}

WORKBOOK_KEYWORDS = {
    "assumptions": ["input", "assumption", "story", "driver"],
    "valuation": ["valuation", "output", "summary", "picture", "football", "bridge"],
    "cost_of_capital": ["cost of capital", "wacc", "rating", "beta"],
    "adjustments": ["r&d", "lease", "normalize", "option"],
    "diagnostics": ["diagnostic", "check", "sensitivity", "scenario"],
}

BUFFETT_SCAN_PAGE_LIMIT = 250
WORKBOOK_FORMULA_SAMPLE_LIMIT = 20000

HEADING_RE = re.compile(r"^[A-Z][A-Z0-9&/()'.,\- ]{4,}$")
TICKER_RE = re.compile(r"\(([A-Z]{1,5}(?:-[A-Z]{2,4})?)\)")
ACTION_WORDS = ("exit", "exited", "trim", "trimmed", "added", "add", "initiated", "sold", "bought", "reduced", "increased")
RISK_WORDS = ("risk", "risks", "uncertainty", "headwind", "headwinds", "challenge", "threat")
THESIS_WORDS = ("because", "benefit", "attractive", "cheap", "undervalued", "tailwind", "thesis", "positioned")
VALUATION_WORDS = ("valuation", "multiple", "ev", "ebitda", "earnings", "fcf", "cash flow", "intrinsic value")
POSITION_STOPWORDS = {
    "page",
    "investor letter",
    "quarterly investor letter",
    "annualized net",
    "return",
    "nav",
    "tr",
    "hbm",
    "fund",
    "dear investor",
    "dear clients",
    "conclusion",
    "portfolio updates1",
}


@dataclass
class PdfExtraction:
    page_count: int
    pages: list[str]
    headings: list[str]
    chunks: list[dict[str, Any]]


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def first_match(patterns: Iterable[re.Pattern[str]], text: str) -> str | None:
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return match.group(0)
    return None


def split_paragraphs(text: str) -> list[str]:
    return [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]


def select_paragraphs(paragraphs: Iterable[str], keywords: Iterable[str], limit: int = 6) -> list[str]:
    selected: list[str] = []
    lowered = [kw.lower() for kw in keywords]
    for paragraph in paragraphs:
        hay = paragraph.lower()
        if any(kw in hay for kw in lowered):
            selected.append(paragraph)
        if len(selected) >= limit:
            break
    return selected


def detect_headings(pages: list[str], limit: int = 25) -> list[str]:
    headings: list[str] = []
    for page in pages:
        for line in page.splitlines():
            line = line.strip()
            if not line:
                continue
            if len(line) > 80:
                continue
            lower = line.lower()
            if lower.startswith("page ") or lower in POSITION_STOPWORDS:
                continue
            if HEADING_RE.match(line):
                headings.append(line)
            if len(headings) >= limit:
                return headings
    return headings


def generic_pdf_chunks(pages: list[str], group_size: int = 3) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    for start in range(0, len(pages), group_size):
        end = min(start + group_size, len(pages))
        text = "\n\n".join(pages[start:end]).strip()
        chunks.append(
            {
                "label": f"pages_{start + 1}_{end}",
                "start_page": start + 1,
                "end_page": end,
                "text_preview": text[:1500],
            }
        )
    return chunks


def chunk_buffett_letters(pages: list[str]) -> list[dict[str, Any]]:
    starts: list[tuple[int, str]] = []
    for idx, page in enumerate(pages, start=1):
        lines = [line.strip() for line in page.splitlines() if line.strip()]
        if not lines:
            continue
        head = lines[0]
        if re.match(r"^\d{4}\s+Letter$", head):
            starts.append((idx, head))
    chunks: list[dict[str, Any]] = []
    for i, (start_page, label) in enumerate(starts[:40]):
        end_page = (starts[i + 1][0] - 1) if i + 1 < len(starts) else len(pages)
        sample_text = "\n\n".join(pages[start_page - 1:min(end_page, start_page + 2)])
        chunks.append(
            {
                "label": label,
                "start_page": start_page,
                "end_page": end_page,
                "text_preview": sample_text[:2000],
            }
        )
    return chunks


def extract_buffett_pdf(path: Path) -> PdfExtraction:
    reader = PdfReader(str(path))
    pages: list[str] = []
    for idx, page in enumerate(reader.pages, start=1):
        if idx > BUFFETT_SCAN_PAGE_LIMIT:
            break
        pages.append(clean_text(page.extract_text() or ""))
    headings = detect_headings(pages)
    chunks = chunk_buffett_letters(pages)
    return PdfExtraction(page_count=len(reader.pages), pages=pages, headings=headings, chunks=chunks)


def extract_pdf(path: Path, record_id: str) -> PdfExtraction:
    if record_id == "buffett_letters_compilation":
        return extract_buffett_pdf(path)
    reader = PdfReader(str(path))
    pages = [clean_text(page.extract_text() or "") for page in reader.pages]
    headings = detect_headings(pages)
    group_size = 2 if len(pages) <= 12 else 3
    chunks = generic_pdf_chunks(pages, group_size=group_size)
    return PdfExtraction(page_count=len(pages), pages=pages, headings=headings, chunks=chunks)


def is_plausible_person_name(value: str) -> bool:
    tokens = value.strip().split()
    if not 2 <= len(tokens) <= 5:
        return False
    if len(value) > 60:
        return False
    if any(token.lower() in POSITION_STOPWORDS for token in tokens):
        return False
    return all(re.fullmatch(r"[A-Z][A-Za-z.'-]*", token) for token in tokens)


def detect_manager(text: str) -> str | None:
    lines = [line.strip() for line in text.splitlines()]
    for idx, line in enumerate(lines):
        if line.lower().startswith("from:"):
            candidate = line.split(":", 1)[1].strip()
            if is_plausible_person_name(candidate):
                return candidate
        if line.lower().startswith("chief executive officer:"):
            candidate = line.split(":", 1)[1].strip()
            if is_plausible_person_name(candidate):
                return candidate
        if line.lower() in {"sincerely,", "yours,", "yours faithfully,"}:
            for follow in lines[idx + 1 : idx + 5]:
                if is_plausible_person_name(follow):
                    return follow
    return None


def infer_strategy_style(text: str) -> str:
    hay = text.lower()
    tags: list[str] = []
    if "short" in hay and "long" in hay:
        tags.append("long/short")
    if "value" in hay:
        tags.append("value-oriented")
    if "hedge" in hay:
        tags.append("hedge-fund style")
    if "small-cap" in hay or "microcap" in hay:
        tags.append("small-cap")
    if "partnership" in hay:
        tags.append("partnership letter")
    if not tags:
        tags.append("public-equity investor")
    return ", ".join(dict.fromkeys(tags))


def extract_positions(paragraphs: list[str], limit: int = 10) -> list[str]:
    positions: list[str] = []
    seen: set[str] = set()
    name_pattern = re.compile(r"^([A-Z][A-Za-z0-9&.'\- ]{2,60})(?:\s+\([A-Z]{1,6}(?:-[A-Z]{2,4})?\))?\s*$")
    for paragraph in paragraphs:
        first_line = paragraph.splitlines()[0].strip()
        match = name_pattern.match(first_line)
        if match:
            name = match.group(1).strip()
            normalized = name.lower()
            if (
                normalized not in seen
                and len(name.split()) <= 6
                and not any(stop in normalized for stop in POSITION_STOPWORDS)
                and not re.fullmatch(r"[A-Z]{1,4}", name)
                and not any(char.isdigit() for char in name)
            ):
                seen.add(name.lower())
                positions.append(name)
        for ticker_match in TICKER_RE.finditer(paragraph):
            ticker = ticker_match.group(1)
            normalized_ticker = ticker.lower()
            if normalized_ticker not in seen and normalized_ticker not in POSITION_STOPWORDS:
                seen.add(ticker.lower())
                positions.append(ticker)
        if len(positions) >= limit:
            break
    return positions[:limit]


def summarize_paragraphs(paragraphs: list[str], limit: int = 5, max_len: int = 280) -> list[str]:
    result: list[str] = []
    for paragraph in paragraphs[:limit]:
        compact = re.sub(r"\s+", " ", paragraph)
        result.append(compact[:max_len])
    return result


def extract_quotes(text: str, limit: int = 4) -> list[str]:
    quotes = re.findall(r'“([^”]{20,240})”|"([^"\n]{20,240})"', text)
    flattened: list[str] = []
    for pair in quotes:
        value = next((item for item in pair if item), "")
        if value:
            flattened.append(value.strip())
        if len(flattened) >= limit:
            break
    return flattened


def extract_investor_letter_fields(record: Any, pdf: PdfExtraction) -> dict[str, Any]:
    text = "\n\n".join(pdf.pages[: min(len(pdf.pages), 12)])
    front_text = "\n\n".join(pdf.pages[: min(len(pdf.pages), 2)])
    paragraphs = split_paragraphs(text)
    headings = pdf.headings
    performance = select_paragraphs(paragraphs, ["return", "%", "benchmark", "performance", "since inception"], limit=3)
    valuation = select_paragraphs(paragraphs, VALUATION_WORDS, limit=5)
    risks = select_paragraphs(paragraphs, RISK_WORDS, limit=5)
    actions = select_paragraphs(paragraphs, ACTION_WORDS, limit=5)
    thesis = select_paragraphs(paragraphs, THESIS_WORDS, limit=6)
    return {
        "title": record.title,
        "manager": "Warren Buffett" if record.id == "buffett_letters_compilation" else detect_manager(front_text),
        "date": first_match(DATE_PATTERNS, front_text),
        "strategy_style": infer_strategy_style(text),
        "performance_context": summarize_paragraphs(performance),
        "top_themes": headings[:8],
        "positions_discussed": extract_positions(paragraphs),
        "thesis_points": summarize_paragraphs(thesis),
        "valuation_language": summarize_paragraphs(valuation),
        "risks_discussed": summarize_paragraphs(risks),
        "portfolio_actions": summarize_paragraphs(actions),
        "tone_notes": record.summary,
        "best_quotes_or_phrases": extract_quotes(text),
    }


def extract_investor_presentation_fields(record: Any, pdf: PdfExtraction) -> dict[str, Any]:
    text = "\n\n".join(pdf.pages[: min(len(pdf.pages), 15)])
    paragraphs = split_paragraphs(text)
    return {
        "title": record.title,
        "manager": detect_manager(text),
        "date": first_match(DATE_PATTERNS, text),
        "agenda": pdf.headings[:10],
        "portfolio_contributors": summarize_paragraphs(select_paragraphs(paragraphs, ["contributor", "top contributor"], limit=4)),
        "portfolio_detractors": summarize_paragraphs(select_paragraphs(paragraphs, ["detractor", "top detractor"], limit=4)),
        "holdings_covered": extract_positions(paragraphs, limit=12),
        "valuation_points": summarize_paragraphs(select_paragraphs(paragraphs, VALUATION_WORDS, limit=6)),
        "exit_or_trim_logic": summarize_paragraphs(select_paragraphs(paragraphs, ["exit", "trim", "sold", "reduced"], limit=4)),
        "slide_structure_patterns": pdf.headings[:12],
    }


def extract_case_study_fields(record: Any, pdf: PdfExtraction) -> dict[str, Any]:
    text = "\n\n".join(pdf.pages[: min(len(pdf.pages), 12)])
    paragraphs = split_paragraphs(text)
    methods = []
    for method in ["DCF", "NPV", "IRR", "comparable companies", "comparable transactions", "LBO"]:
        if method.lower() in text.lower():
            methods.append(method)
    return {
        "title": record.title,
        "case_type": "valuation case study",
        "company_or_asset": record.title,
        "task_prompt": summarize_paragraphs(select_paragraphs(paragraphs, ["task", "question", "asked", "recommendation"], limit=4)),
        "core_assumptions": summarize_paragraphs(select_paragraphs(paragraphs, ["assumption", "margin", "growth", "cost of capital", "working capital"], limit=6)),
        "expected_outputs": methods,
        "valuation_methods": methods,
        "decision_required": summarize_paragraphs(select_paragraphs(paragraphs, ["recommend", "decision", "should", "sell", "invest"], limit=4)),
        "grading_notes": [record.summary],
    }


def count_keywords(names: list[str], keywords: dict[str, list[str]]) -> dict[str, list[str]]:
    matched: dict[str, list[str]] = {key: [] for key in keywords}
    for name in names:
        lower = name.lower()
        for category, tokens in keywords.items():
            if any(token in lower for token in tokens):
                matched[category].append(name)
    return matched


def extract_xlsx_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(filename=path, data_only=False, read_only=True)
    sheet_names = wb.sheetnames
    formula_counts: dict[str, int] = {}
    dimensions: dict[str, str] = {}
    sampled_cells: dict[str, int] = {}
    for name in sheet_names:
        ws = wb[name]
        count = 0
        seen = 0
        for row in ws.iter_rows():
            for cell in row:
                seen += 1
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    count += 1
                if seen >= WORKBOOK_FORMULA_SAMPLE_LIMIT:
                    break
            if seen >= WORKBOOK_FORMULA_SAMPLE_LIMIT:
                break
        formula_counts[name] = count
        dimensions[name] = ws.calculate_dimension()
        sampled_cells[name] = seen
    return {"sheet_names": sheet_names, "formula_counts": formula_counts, "dimensions": dimensions, "sampled_cells": sampled_cells}


def extract_xls_workbook(path: Path) -> dict[str, Any]:
    wb = xlrd.open_workbook(path)
    sheet_names = wb.sheet_names()
    dimensions: dict[str, str] = {}
    formula_counts: dict[str, int] = {}
    sampled_cells: dict[str, int] = {}
    for name in sheet_names:
        sh = wb.sheet_by_name(name)
        dimensions[name] = f"A1:{sh.nrows}x{sh.ncols}"
        formula_counts[name] = 0
        sampled_cells[name] = min(sh.nrows * sh.ncols, WORKBOOK_FORMULA_SAMPLE_LIMIT)
    return {"sheet_names": sheet_names, "formula_counts": formula_counts, "dimensions": dimensions, "sampled_cells": sampled_cells}


def infer_model_type(title: str, sheet_names: list[str]) -> str:
    hay = " ".join([title, *sheet_names]).lower()
    if "comparable" in hay or "comps" in hay:
        return "comparable company analysis"
    if "lbo" in hay:
        return "lbo"
    if "fcff" in hay or "dcf" in hay:
        return "fcff / dcf"
    if "synergy" in hay:
        return "merger synergy calculator"
    return "financial model workbook"


def extract_workbook_fields(record: Any, path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        extracted = extract_xlsx_workbook(path)
    else:
        extracted = extract_xls_workbook(path)
    sheet_names = extracted["sheet_names"]
    matched = count_keywords(sheet_names, WORKBOOK_KEYWORDS)
    top_formula_sheets = sorted(extracted["formula_counts"].items(), key=lambda item: (-item[1], item[0]))[:8]
    formula_summary = [
        f"{name}: {count} formulas in first {extracted['sampled_cells'][name]} sampled cells"
        for name, count in top_formula_sheets
        if count > 0
    ]
    if not formula_summary:
        formula_summary = [f"{name}: structure only" for name in sheet_names[:5]]
    quality_notes = []
    if any("answer" in name.lower() for name in sheet_names):
        quality_notes.append("Contains answer-key or teaching-oriented tabs.")
    if any("read me" in name.lower() for name in sheet_names):
        quality_notes.append("Includes explicit instructional/readme tab, indicating template usage.")
    return (
        {
            "title": record.title,
            "model_type": infer_model_type(record.title, sheet_names),
            "sheet_names": sheet_names,
            "key_schedules": matched["assumptions"] + matched["cost_of_capital"] + matched["adjustments"],
            "core_formulas": formula_summary,
            "assumption_categories": matched["assumptions"] + matched["cost_of_capital"],
            "output_tabs": matched["valuation"],
            "sensitivity_support": matched["diagnostics"],
            "quality_notes": quality_notes or [record.summary],
        },
        extracted,
    )


def extract_generic_fields(record: Any, pdf: PdfExtraction) -> dict[str, Any]:
    text = "\n\n".join(pdf.pages[: min(len(pdf.pages), 10)])
    paragraphs = split_paragraphs(text)
    return {
        "title": record.title,
        "summary": record.summary,
        "notes": summarize_paragraphs(paragraphs, limit=5),
    }


def apply_manual_override(stub: dict[str, Any], record: Any) -> dict[str, Any]:
    fields = dict(stub["fields"])
    fields.update(MANUAL_OVERRIDES[record.id])
    stub["fields"] = fields
    stub["extraction_status"] = "extracted"
    stub["source_metadata"] = {
        "extraction_mode": "manual_override",
        "file_present": False,
    }
    stub["chunking"] = {
        "strategy": "manual_curated",
        "chunks": [],
    }
    stub["text_samples"] = []
    return stub


def populate_stub(stub: dict[str, Any], record: Any) -> dict[str, Any]:
    if record.id in MANUAL_OVERRIDES:
        return apply_manual_override(stub, record)

    path = Path(record.file_path)
    if not path.exists():
        stub["extraction_status"] = "missing_source"
        stub["source_metadata"] = {"file_present": False, "extraction_mode": "missing_source"}
        return stub

    metadata: dict[str, Any] = {
        "file_present": True,
        "source_extension": path.suffix.lower(),
    }

    if path.suffix.lower() == ".pdf":
        pdf = extract_pdf(path, record.id)
        metadata.update(
            {
                "extraction_mode": "pdf_text",
                "page_count": pdf.page_count,
                "headings_detected": pdf.headings[:20],
            }
        )
        if record.template_key == "investor_letter":
            fields = extract_investor_letter_fields(record, pdf)
        elif record.template_key == "investor_presentation":
            fields = extract_investor_presentation_fields(record, pdf)
        elif record.template_key == "case_study":
            fields = extract_case_study_fields(record, pdf)
        else:
            fields = extract_generic_fields(record, pdf)
        chunking = {
            "strategy": "letter_year_chunks" if record.id == "buffett_letters_compilation" else "page_groups",
            "chunk_count": len(pdf.chunks),
            "chunks": pdf.chunks[:40],
        }
        text_samples = [chunk["text_preview"] for chunk in pdf.chunks[:5]]
    elif path.suffix.lower() in {".xlsx", ".xls"}:
        fields, workbook_meta = extract_workbook_fields(record, path)
        metadata.update(
            {
                "extraction_mode": "workbook_structure",
                **workbook_meta,
            }
        )
        chunking = {
            "strategy": "sheet_level",
            "chunk_count": len(workbook_meta["sheet_names"]),
            "chunks": [
                {
                    "label": name,
                    "start_page": None,
                    "end_page": None,
                    "text_preview": f"dimension={workbook_meta['dimensions'][name]}; formulas={workbook_meta['formula_counts'][name]}",
                }
                for name in workbook_meta["sheet_names"][:40]
            ],
        }
        text_samples = [chunk["text_preview"] for chunk in chunking["chunks"][:8]]
    else:
        stub["extraction_status"] = "unsupported_source"
        stub["source_metadata"] = {"file_present": True, "extraction_mode": "unsupported_source"}
        return stub

    merged_fields = dict(stub["fields"])
    merged_fields.update(fields)
    stub["fields"] = merged_fields
    stub["source_metadata"] = metadata
    stub["chunking"] = chunking
    stub["text_samples"] = text_samples
    stub["extraction_status"] = "extracted"
    stub["extracted_at"] = datetime.now(timezone.utc).isoformat()
    return stub


def main() -> None:
    parser = argparse.ArgumentParser(description="Populate extracted source JSON files from registry entries.")
    parser.add_argument("--phase", type=int, default=1, help="Only process this recommended phase.")
    parser.add_argument("--bucket", type=str, default=None, help="Optional primary bucket filter.")
    parser.add_argument("--include-excluded", action="store_true", help="Include concept-only / excluded records.")
    args = parser.parse_args()

    records = filter_sources(
        load_registry(),
        phase=args.phase,
        bucket=args.bucket,
        include_excluded=args.include_excluded,
    )
    templates = load_templates()

    target_dir = OUT_DIR / f"phase_{args.phase}"
    if args.bucket:
        target_dir = target_dir / args.bucket
    target_dir.mkdir(parents=True, exist_ok=True)

    manifest: list[dict[str, Any]] = []
    status_counter: Counter[str] = Counter()
    for record in records:
        stub = build_stub(record, templates)
        populated = populate_stub(stub, record)
        out_path = target_dir / f"{record.id}.json"
        out_path.write_text(json.dumps(populated, indent=2) + "\n", encoding="utf-8")
        status = populated.get("extraction_status", "unknown")
        status_counter[status] += 1
        manifest.append(
            {
                "id": record.id,
                "path": str(out_path),
                "bucket": record.primary_bucket,
                "status": status,
            }
        )

    manifest_path = target_dir / "_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(f"Processed {len(records)} sources into {target_dir}")
    print(f"Manifest: {manifest_path}")
    print(json.dumps(status_counter, indent=2))


if __name__ == "__main__":
    main()
